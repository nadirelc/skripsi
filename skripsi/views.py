from django.http import HttpResponse
from django.shortcuts import render
from django.template import loader
from django.shortcuts import redirect
from django.conf import settings
from django.core.files.storage import FileSystemStorage
from django.http import FileResponse, Http404
import os

#IMPORT REF SKRIPSI NADIR
from refextract import extract_references_from_file
from refextract import extract_references_from_url
from gensim.parsing.preprocessing import remove_stopwords
import xlsxwriter
import datetime
import unidecode
import openpyxl
import re
import slate3k
import urllib.request


def index(request):
    return render(request,'form.html')

def refForm(request):
    #DATA SINGKATAN

    #END DATA SINGKATAN
    abv_dict = {}
    file_loc = os.path.join(os.path.dirname(os.path.dirname(__file__)),'media/excel/abbreviation_dictionary.xlsx')
    wb_obj = openpyxl.load_workbook(file_loc)
    sheet_obj = wb_obj.active
    m_row = sheet_obj.max_row
    for i in range(1, m_row + 1):
        cell_obj_abv = sheet_obj.cell(row = i, column = 1)
        cell_obj_abv2 = sheet_obj.cell(row = i, column = 2)
        abv_dict[cell_obj_abv.value] = cell_obj_abv2.value
    #VAR WORLD

    #DATA KOTA
    file_city =  os.path.join(os.path.dirname(os.path.dirname(__file__)),'media/excel/worldcities.xlsx')
    wb_obj = openpyxl.load_workbook(file_city)
    sheet_obj = wb_obj.active
    cell_obj = sheet_obj.cell(row = 1, column = 1)
    m_row = sheet_obj.max_row

    array_city = []
    for i in range(1, m_row + 1):
        cell_obj = sheet_obj.cell(row = i, column = 1)
        array_city.append(cell_obj.value)

    #DATA SCOPUS
    file_city2 =  os.path.join(os.path.dirname(os.path.dirname(__file__)),'media/excel/scopus_journal.xlsx')
    wb_obj2 = openpyxl.load_workbook(file_city2)
    sheet_obj2 = wb_obj2.active
    cell_obj2 = sheet_obj2.cell(row = 1, column = 1)
    m_row2 = sheet_obj2.max_row
    array_journal = []
    for i in range(1, m_row2 + 1):
        cell_obj2 = sheet_obj2.cell(row = i, column = 1)
        array_journal.append(cell_obj2.value)

    #DATA BLACK LIST PREDATORY JOURNAL 
    file_city4 =  os.path.join(os.path.dirname(os.path.dirname(__file__)),'media/excel/predatoryjurnal_abv.xlsx')
    wb_obj4 = openpyxl.load_workbook(file_city4)
    sheet_obj4 = wb_obj4.active
    cell_obj4 = sheet_obj4.cell(row = 1, column = 1)
    m_row4 = sheet_obj4.max_row
    array_predatory = []
    for i in range(1, m_row4 + 1):
        cell_obj4 = sheet_obj4.cell(row = i, column = 1)
        array_predatory.append(cell_obj4.value)
    
    #DATA BLACK LIST PREDATORY PUBLISHER
    file_city5 =  os.path.join(os.path.dirname(os.path.dirname(__file__)),'media/excel/predatorypublisher_abv.xlsx')
    wb_obj5 = openpyxl.load_workbook(file_city5)
    sheet_obj5 = wb_obj5.active
    cell_obj5 = sheet_obj5.cell(row = 1, column = 1)
    m_row5 = sheet_obj5.max_row
    for i in range(1, m_row5 + 1):
        cell_obj5 = sheet_obj5.cell(row = i, column = 1)
        array_predatory.append(cell_obj5.value)

    #DATA SCOPUS CONF
    file_city3 =  os.path.join(os.path.dirname(os.path.dirname(__file__)),'media/excel/scopus_conference.xlsx')
    wb_obj3 = openpyxl.load_workbook(file_city3)
    sheet_obj3 = wb_obj3.active
    cell_obj3 = sheet_obj3.cell(row = 1, column = 1)
    m_row3 = sheet_obj3.max_row
    array_conf = []
    for i in range(1, m_row3 + 1):
        cell_obj3 = sheet_obj3.cell(row = i, column = 1)
        array_conf.append(cell_obj3.value)
        
    #END VAR WORLD

    #FUNCTION FOR SYSTEM

    def cekJurnalIlmiah(ref):
        detail_format = []
        if 'vol. ' not in listToString(ref['raw_ref']).lower():
            detail_format.append("*V")
        if 'no. ' not in listToString(ref['raw_ref']).lower():
            detail_format.append("*N")
        if 'p. ' not in listToString(ref['raw_ref']).lower():
            detail_format.append("*P")
        if 'year' not in ref:
            detail_format.append("*T")
        
        ####### CEK APAKAH ADA NAMA JURNAL APDA REFERENSI ##########
        ref_judul = listToString(ref['raw_ref'])
        #UNTUJK YANG MENGGUNAKAN PETIK MIRING
        ref_judul2 = ref_judul.split('“')
        if len(ref_judul2) > 1:
            ref_judul2 = ref_judul2[1].split('”')
            if len(ref_judul2) > 1:
                ref_judul2 = ref_judul2[1].split(',')
                ref_judul2 = ref_judul2[0]
                ref_judul2 = ref_judul2[1:]
                #print(ref_judul2)
            else:
                detail_format.append("*NJ")
        #UNTUK PETIK BIASA
        else:
            ref_judul2 = ref_judul.split('"')
            if len(ref_judul2) > 2:
                ref_judul2 = ref_judul2[2].split(',')
                ref_judul2 = ref_judul2[0]
                ref_judul2 = ref_judul2[1:]
                #print(ref_judul2)
            else:
                detail_format.append("*NJ")
        
        #print("INI JUDUL 2 YG ERROR = ", ref_judul2)
        #ref_judul2 = remove_stopwords(ref_judul2)
        #array_text = ref_judul2.split()
        #for i in range(len(array_text)):
        #    if array_text[i] in abv_dict:
        #        array_text[i] = abv_dict[array_text[i]]
        #        
        #ref_judul2 = listToString(array_text)

        return detail_format

    def cekConference(ref):
        detail_format = []
        if cekIsBuku(ref):
            # TES TIDAK MENGGUNAKAN Pages
            if 'p. ' in listToString(ref['raw_ref']):
               pass
            else:
               detail_format.append("*P")        
        else:
            detail_format.append("*K")
        
        ####### CEK APAKAH ADA NAMA JURNAL APDA REFERENSI ##########
        ref_judul = listToString(ref['raw_ref'])
        #UNTUJK YANG MENGGUNAKAN PETIK MIRING
        ref_judul2 = ref_judul.split('“')
        if len(ref_judul2) > 1:
            ref_judul2 = ref_judul2[1].split('”')
            if len(ref_judul2) > 1:
                ref_judul2 = ref_judul2[1].split(',')
                ref_judul2 = ref_judul2[0]
                ref_judul2 = ref_judul2[1:]
                #print(ref_judul2)
            else:
                detail_format.append("*NJ")
        #UNTUK PETIK BIASA
        else:
            ref_judul2 = ref_judul.split('"')
            if len(ref_judul2) > 2:
                ref_judul2 = ref_judul2[2].split(',')
                ref_judul2 = ref_judul2[0]
                ref_judul2 = ref_judul2[1:]
                #print(ref_judul2)
            else:
                detail_format.append("*NJ")

        return detail_format
        
    def listToString(s): 
        # initialize an empty string
        str1 = " " 
        # return string  
        return (str1.join(s))

    def listToStringComma(s): 
        str2 = ", "
        return (str2.join(s))

    def cekIsBuku(ref):
        misc = listToString(ref['misc'])
        misc = misc.replace(":", "")
        misc = misc.replace(",", "")
        word = misc.split()
        word_cleared = []
        
        for i in range(len(word)):
            word_cleared.append(unidecode.unidecode(word[i]))

        for kota in word_cleared:
            if kota in array_city:
                return True

        return False 

    def sumberReputasi(ref, jenis):
        ref_judul = listToString(ref['raw_ref'])
        #UNTUJK YANG MENGGUNAKAN PETIK MIRING
        ref_judul2 = ref_judul.split('“')
        
        if len(ref_judul2) > 1:
            ref_judul2 = ref_judul2[1].split('”')
            if len(ref_judul2) > 1:
                ref_judul2 = ref_judul2[1].split(',')
                if(ref_judul2[0] == ''):
                    ref_judul2 = ref_judul2[1]
                else:
                    ref_judul2 = ref_judul2[0]
                ref_judul2 = ref_judul2[1:]
            else:
                return False
        #UNTUK PETIK BIASA
        else:
            ref_judul2 = ref_judul.split('"')
            if len(ref_judul2) > 2:
                ref_judul2 = ref_judul2[2].split(',')
                if(ref_judul2[0] == ''):
                    ref_judul2 = ref_judul2[1]
                else:
                    ref_judul2 = ref_judul2[0]
                ref_judul2 = ref_judul2[1:]
            else:
                return False

        ref_judul2 = remove_stopwords(ref_judul2)
        ref_judul2 = ref_judul2.replace(" &", "")
        array_text = ref_judul2.split()
        for i in range(len(array_text)):
            if array_text[i] in abv_dict:
                array_text[i] = abv_dict[array_text[i]]
                
        ref_judul2 = listToString(array_text)

        if jenis == 'conf':
            array_check = array_conf
        elif jenis == 'predatory':
            array_check = array_predatory
        else:
            array_check = array_journal
                
        #print("Judul akhir = ", ref_judul2)
        if ref_judul2 in array_check:
            #print("ADA BOS SELAMAT KEREN")
            return True
            
        return  False
    
    #END FUCNTION

    file_xpdf = open(os.path.join(settings.BASE_DIR, 'pdftotext.exe'))

    #GETTING THE AUTHOR NAME IN ARRAY
    array_author = []
    for i in range (15):
        name_form = "author_" + str(i+1)
        author = request.POST.get(name_form)
        if author != None:
            if author != '':
                array_author.append(author)
    
    #MENGHILANGKAN AKSEN
    for i in range(len(array_author)):
        array_author[i] = (unidecode.unidecode(array_author[i]))
    
    #Menyingkat nama depan
    for i in range(len(array_author)):
        split_name = array_author[i].split()
        first_name = ""
        for j in range(len(split_name) - 1):
            first_name = first_name + split_name[j][0] + ". "
        array_author[i] = first_name + split_name[len(split_name) - 1]
    
    #GETTING THE PDF FILE by URL
    url_pdf = request.POST.get("url_pdf")

    if url_pdf != '':
        references = extract_references_from_url(url_pdf)
        file_name = ""
        #print("INI URL PDF", url_pdf)
    else:
        file_pdf = request.FILES['file_pdf']
        file = FileSystemStorage()
        file.save(file_pdf.name, file_pdf)
        #print("INI FILE PDF", file_pdf.name)

        file_name = os.path.join(os.path.dirname(os.path.dirname(__file__)),'media/' + file_pdf.name)

        references = extract_references_from_file(file_name)

        #DO WITH FILE

    ## START WRITE EXCEL $$
    #WRITE EXCEL
    workbook = xlsxwriter.Workbook('analisis.xlsx')
    worksheet = workbook.add_worksheet('Output Tabel')
    worksheet.set_column('B:H', 15)
    worksheet.set_column('J:J', 80)

    worksheet_narasi = workbook.add_worksheet('Output Narasi')

    worksheet_narasi.set_column('A:A', 20)
    worksheet_narasi.set_column('B:B', 100)

    cell_format_narasi = workbook.add_format()
    cell_format_narasi.set_bold()
    worksheet_narasi.write('A1', 'OUTPUT NARASI', cell_format_narasi)

    #WRITE TEMPLATE
    worksheet.write('A1', 'Referensi')
    worksheet.write('B1', 'Format')
    worksheet.write('C1', 'Self Citation')
    worksheet.write('D1', 'Acuan Primer')
    worksheet.write('E1', 'Tahun Terbit')
    worksheet.write('F1', 'Tersitasi Pada Paper')
    worksheet.write('G1', 'White List')
    worksheet.write('H1', 'Black List')

    #WRITE TIDAK LENGKAP information
    worksheet.write('J2', '*P = Tidak terdapat Page (p. / pp.)')
    worksheet.write('J3', '*V = Tidak terdapat Volume (vol.)')
    worksheet.write('J4', '*N = Tidak terdapat Nomor jurnal (no.)')
    worksheet.write('J5', '*K = Tidak terdapat Tempat/kota publikasi (Buku/Conference)')
    worksheet.write('J6', '*T = Tidak terdapat Tahun terbit')
    worksheet.write('J7', '*NJ = Tidak terdapat Nama Jurnal')
    worksheet.write('J7', '*TT = Referensi Tidak Teridenfitikasi, sehingga dianggap tidak lengkap')
    
    # WRITE ALL FORMAT TO LENGKAP AND SUMBER REPUTASI TIDAK
    row = 1
    cell_format_sitasi = workbook.add_format()
    cell_format_sitasi.set_bg_color('red')
    cell_format4 = workbook.add_format()
    cell_format4.set_bg_color('yellow')
    for ref in references:
        worksheet.write(row, 1, "Lengkap")
        worksheet.write(row, 5, "Tidak Tersitasi", cell_format_sitasi)
        worksheet.write(row, 6, "Tidak", cell_format4)
        worksheet.write(row, 7, "Tidak", cell_format4)
        row += 1

    #SAVING ARRAY RAW_REF
    raw_ref = []
    for ref in references:
        raw_ref.append(ref['raw_ref'])

    #WRITE ARRAY RAW_REF
    row = 1
    for ref in raw_ref:
        worksheet.write(row, 0, ref[0])
        row += 1

    #SAVING ARRAY year_ref

    year_ref = []
    for ref in references:
        if 'year' in ref:
            year_ref.append(ref['year'])
        else:
            year_ref.append(["Kosong"])


    #WRITE ARRAY year_ref
    row = 1
    date_now = datetime.date.today()
    year_now = date_now.strftime("%Y")
    year_red = 0
    year_red_array = []
    total_year = len(year_ref)
    for i, ref in enumerate(year_ref):
        cell_format = workbook.add_format()
        if ref[0] == "Kosong":
            year_red += 1
            year_red_array.append(str(i+1))
            cell_format.set_bg_color('red')
            worksheet.write(row, 1, "Tidak Lengkap (*T)", cell_format)
        elif (int(year_now) - int(ref[0])) > 10:
            year_red += 1
            year_red_array.append(str(i+1))
            cell_format.set_bg_color('red')
        
        worksheet.write(row, 4, ref[0], cell_format)
        row += 1

    #BORDER ALL NARASI
    cell_format_primer2 = workbook.add_format()
    cell_format_primer2.set_border()
    cell_format_primer2.set_text_wrap()

    #WRITE YEAR IN OUTPUT NARASI
    year_percentage = year_red * 100 / total_year
    cell_format_year = workbook.add_format()
    cell_format_year.set_text_wrap()
    cell_format_year.set_border()
    year_percentage_str = "%.2f" % year_percentage
    year_percentage_str2 = 100 - year_percentage
    year_percentage_str2 = "%.2f" % year_percentage_str2
    if year_percentage > 20:
        cell_format_year.set_bg_color('red')
        narasi_year = 'Referensi yang digunakan pada paper ini dianggap mutakhir sebesar ' + str(year_percentage_str2) + '%. Referensi dianggap mutakhir ketika tahun terbit referensi berumur maksimal 10 tahun pada saat sistem digunakan. \n \nReferensi dengan Nomor ' + listToStringComma(year_red_array) + ' tidak mutakhir.'
        worksheet_narasi.write('A4', 'Tahun terbit', cell_format_year)
        worksheet_narasi.write('B4', narasi_year, cell_format_primer2)
    else:
        cell_format_year.set_bg_color('#00ac56')
        if (len(year_red_array) > 0):
            narasi_year = 'Referensi yang digunakan pada paper ini dianggap mutakhir sebesar ' + str(year_percentage_str2) + '%. Referensi dianggap mutakhir ketika tahun terbit referensi berumur maksimal 10 tahun pada saat sistem digunakan. \n \nReferensi dengan Nomor ' + listToStringComma(year_red_array) + ' tidak mutakhir.'
        else:
            narasi_year = 'Referensi yang digunakan pada paper ini dianggap mutakhir sebesar ' + str(year_percentage_str2) + '%. Referensi dianggap mutakhir ketika tahun terbit referensi berumur maksimal 10 tahun pada saat sistem digunakan.'
        worksheet_narasi.write('A4', 'Tahun terbit', cell_format_year)
        worksheet_narasi.write('B4', narasi_year, cell_format_primer2)


    #SAVING Array of Author
    ref_author = []
    row = 1
    cell_format3 = workbook.add_format()
    cell_format3.set_bg_color('red')
    for ref in references:
        if 'author' in ref:
            ref_author.append(ref['author'])
        else:
            ref_author.append(["Tidak ada author"])
            #TIDAK ADA AUTHOR TIDAK LENGKAP
            #worksheet.write(row, 1, "Tidak Lengkap", cell_format3)
        row += 1
    for i in range(len(ref_author)):
        ref_author[i] = (unidecode.unidecode(ref_author[i][0]))
    #Write Self Citation Tidak
    row = 1
    for ref in references:
        worksheet.write(row, 2, "Tidak")
        row += 1

    #Write Self Citation IYA
    row = 1
    self_citation = 0
    self_citation_array = []
    self_citation_total = len(ref_author)

    for i, ref in enumerate(ref_author):
        cell_format = workbook.add_format()
        cell_format.set_bg_color('yellow')
        for author_name in array_author:
            if author_name in ref:
                self_citation += 1
                self_citation_array.append(str(i+1))
                worksheet.write(row, 2, "Iya", cell_format)
        row += 1
    
    #WRITE SELF CITATION IN OUTPUT NARASI
    self_citation_percentage = self_citation * 100 / self_citation_total
    cell_format_citation = workbook.add_format()
    cell_format_citation.set_border()
    cell_format_citation.set_text_wrap()
    citation_percentage_str = "%.2f" % self_citation_percentage
    citation_percentage_str2 = 100 - self_citation_percentage
    citation_percentage_str2 = "%.2f" % citation_percentage_str2 
    if self_citation_percentage > 10:
        cell_format_citation.set_bg_color('red')
        narasi_citation = 'Referensi yang digunakan pada paper ini terindikasi self-citation sebesar ' + str(citation_percentage_str) + '%. Sebuah referensi dianggap self-citation jika nama penulis paper dan nama author pada referensi memiliki kesamaan. \n \nReferensi dengan Nomor ' + listToStringComma(self_citation_array) +' terindikasi self-citation.'
        worksheet_narasi.write('A5', 'Self Citation', cell_format_citation)
        worksheet_narasi.write('B5', narasi_citation, cell_format_primer2)
    else:
        cell_format_citation.set_bg_color('#00ac56')
        if (len(self_citation_array) > 0):
            narasi_citation = 'Referensi yang digunakan pada paper ini terindikasi self-citation sebesar ' + str(citation_percentage_str) + '%. Sebuah referensi dianggap self-citation jika nama penulis paper dan nama author pada referensi memiliki kesamaan.  \n \nReferensi dengan Nomor ' + listToStringComma(self_citation_array) +' terindikasi self-citation.'
        else:
            narasi_citation = 'Referensi yang digunakan pada paper ini terindikasi self-citation sebesar ' + str(citation_percentage_str) + '%. Sebuah referensi dianggap self-citation jika nama penulis paper dan nama author pada referensi memiliki kesamaan.'
        worksheet_narasi.write('A5', 'Self Citation', cell_format_citation)
        worksheet_narasi.write('B5', narasi_citation, cell_format_primer2)

    #Write Acuan Primer Mungkin Tidak
    row = 1
    cell_format = workbook.add_format()
    cell_format2 = workbook.add_format()
    cell_format.set_bg_color('yellow')
    for ref in references:
        worksheet.write(row, 3, "Tidak", cell_format)
        row += 1

    #KELENGKAPAN SITASI
    not_citated = []
    len_ref = len(references)
    len_ref = int(references[len_ref-1]['linemarker'][0])
    string_inside_brackets = []

    if url_pdf != '':
        req = urllib.request.Request(url_pdf, headers={'User-Agent' : "Magic Browser"})
        #SUDAH TIDAK DIGUNAKAN, HANYA MENERIMA INPUT FILE
        #remote_file = urllib.request.urlopen(req).read()
        #pdf_citation = io.BytesIO(remote_file)
        #pdfReader = PyPDF2.PdfFileReader(pdf_citation, strict=False)
    else:
        with open(file_name,'rb') as f:
            extracted_text = slate3k.PDF(f)

    object_string = str(extracted_text)

    s_filter  = ' '.join(re.findall('\[([^a-z^A-Z]+)\]', object_string))
    s_filter = s_filter.replace("\n", " ")
    s_filter = s_filter.replace(",", " ")
    s_filter = s_filter.replace("[", " ")
    s_filter = s_filter.replace("]", " ")
    array_text_sitasi = s_filter.split(' ')

    for val_array_text in array_text_sitasi:
        if '-' in val_array_text:
            try:
                res = sum(((list(range(*[int(b) + c  
                    for c, b in enumerate(a.split('-'))]))
                    if '-' in a else [int(a)]) for a in val_array_text.split(', ')), [])
                string_inside_brackets += res
            except:
                print("Masuk Except")

    array_text_sitasi = array_text_sitasi[:-len_ref]

    for val_string in string_inside_brackets:
        array_text_sitasi.append(str(val_string))
        
    row = 1
    for ref in references:
        if ref['linemarker'][0] in array_text_sitasi:
            worksheet.write(row, 5, "Tersitasi")
        else :
            not_citated.append(ref['linemarker'][0])
        row += 1
    
    # END KELENGKAPAN SITASI

    #Write Acuan Primer & Format & Sumber reputasi

    #GET DATA PRIMER
    primer = 0
    primer_array = []
    primer_total = len(references)

    #GET DATA FORMAT
    format_lengkap = 0
    format_lengkap_array = []
    format_total = len(references)

    #SUMBER BEREPUTASI
    reputasi = 0
    reputasi_array = []
    reputasi_total = len(references)

    #PREDATORY
    predatory = 0
    predatory_array = []
    predatory_total = len(references)

    row = 1
    for i, ref in enumerate(references):
        #JURNAL ILMIAH
        if 'journal_title' in ref:
            detail_format = cekJurnalIlmiah(ref)
            if detail_format:
                worksheet.write(row, 1, "Tidak Lengkap(" + listToString(detail_format) + ")", cell_format3)
                format_lengkap += 1
                format_lengkap_array.append(str(i+1))
            worksheet.write(row, 3, "Iya")
            primer += 1
            primer_array.append(str(i+1))
            if sumberReputasi(ref, 'jurnal'):
                worksheet.write(row, 6, "Iya")
                reputasi += 1
                reputasi_array.append(str(i+1))
            if sumberReputasi(ref, 'predatory'):
                worksheet.write(row, 7, "Iya")
                predatory += 1
                predatory_array.append(str(i+1))
        elif 'journal_volume' in ref:
            detail_format = cekJurnalIlmiah(ref)
            if detail_format:
                worksheet.write(row, 1, "Tidak Lengkap(" + listToString(detail_format) + ")", cell_format3)
                format_lengkap += 1
                format_lengkap_array.append(str(i+1))
            worksheet.write(row, 3, "Iya")
            primer += 1
            primer_array.append(str(i+1))
            if sumberReputasi(ref, 'jurnal'):
                worksheet.write(row, 6, "Iya")
                reputasi += 1
                reputasi_array.append(str(i+1))
            if sumberReputasi(ref, 'predatory'):
                worksheet.write(row, 7, "Iya")
                predatory += 1
                predatory_array.append(str(i+1))
        elif 'journal_year' in ref:
            detail_format = cekJurnalIlmiah(ref)
            if detail_format:
                worksheet.write(row, 1, "Tidak Lengkap(" + listToString(detail_format) + ")", cell_format3)
                format_lengkap += 1
                format_lengkap_array.append(str(i+1))
            worksheet.write(row, 3, "Iya")
            primer += 1
            primer_array.append(str(i+1))
            if sumberReputasi(ref, 'jurnal'):
                worksheet.write(row, 6, "Iya")
                reputasi += 1
                reputasi_array.append(str(i+1))
            if sumberReputasi(ref, 'predatory'):
                worksheet.write(row, 7, "Iya")
                predatory += 1
                predatory_array.append(str(i+1))
        elif 'journal_page' in ref:
            detail_format = cekJurnalIlmiah(ref)
            if detail_format:
                worksheet.write(row, 1, "Tidak Lengkap(" + listToString(detail_format) + ")", cell_format3)
                format_lengkap += 1
                format_lengkap_array.append(str(i+1))
            worksheet.write(row, 3, "Iya")
            primer += 1
            primer_array.append(str(i+1))
            if sumberReputasi(ref, 'jurnal'):
                worksheet.write(row, 6, "Iya")
                reputasi += 1
                reputasi_array.append(str(i+1))
            if sumberReputasi(ref, 'predatory'):
                worksheet.write(row, 7, "Iya")
                predatory += 1
                predatory_array.append(str(i+1))
        elif 'journal_references' in ref:
            detail_format = cekJurnalIlmiah(ref)
            if detail_format:
                worksheet.write(row, 1, "Tidak Lengkap(" + listToString(detail_format) + ")", cell_format3)
                format_lengkap += 1
                format_lengkap_array.append(str(i+1))
            worksheet.write(row, 3, "Iya")
            primer += 1
            primer_array.append(str(i+1))
            if sumberReputasi(ref, 'jurnal'):
                worksheet.write(row, 6, "Iya")
                reputasi += 1
                reputasi_array.append(str(i+1))
            if sumberReputasi(ref, 'predatory'):
                worksheet.write(row, 7, "Iya")
                predatory += 1
                predatory_array.append(str(i+1))
        elif 'misc' in ref:
            if 'Journal' in listToString(ref['misc']):
                detail_format = cekJurnalIlmiah(ref)
                if detail_format:
                    worksheet.write(row, 1, "Tidak Lengkap(" + listToString(detail_format) + ")", cell_format3)
                    format_lengkap += 1
                    format_lengkap_array.append(str(i+1))
                worksheet.write(row, 3, "Iya")
                primer += 1
                primer_array.append(str(i+1))
                if sumberReputasi(ref, 'jurnal'):
                    worksheet.write(row, 6, "Iya")
                    reputasi += 1
                    reputasi_array.append(str(i+1))
                if sumberReputasi(ref, 'predatory'):
                    worksheet.write(row, 7, "Iya")
                    predatory += 1
                    predatory_array.append(str(i+1))
            elif 'Jurnal' in listToString(ref['misc']):
                detail_format = cekJurnalIlmiah(ref)
                if detail_format:
                    worksheet.write(row, 1, "Tidak Lengkap(" + listToString(detail_format) + ")", cell_format3)
                    format_lengkap += 1
                    format_lengkap_array.append(str(i+1))
                worksheet.write(row, 3, "Iya")
                primer += 1
                primer_array.append(str(i+1))
                if sumberReputasi(ref, 'jurnal'):
                    worksheet.write(row, 6, "Iya")
                    reputasi += 1
                    reputasi_array.append(str(i+1))
                if sumberReputasi(ref, 'predatory'):
                    worksheet.write(row, 7, "Iya")
                    predatory += 1
                    predatory_array.append(str(i+1))
            #CONFERENCE & PROCEEDINGS
            elif 'Conference' in listToString(ref['misc']):
                primer += 1
                primer_array.append(str(i+1))
                worksheet.write(row, 3, "Iya")
                detail_format = cekConference(ref)
                if detail_format:
                    worksheet.write(row, 1, "Tidak Lengkap(" + listToString(detail_format) + ")", cell_format3)
                    format_lengkap += 1
                    format_lengkap_array.append(str(i+1))
                if sumberReputasi(ref, 'conf'):
                    worksheet.write(row, 6, "Iya")
                    reputasi += 1
                    reputasi_array.append(str(i+1))
            elif 'Proceeding' in listToString(ref['misc']):
                primer += 1
                primer_array.append(str(i+1))
                worksheet.write(row, 3, "Iya")
                detail_format = cekConference(ref)
                if detail_format:
                    worksheet.write(row, 1, "Tidak Lengkap(" + listToString(detail_format) + ")", cell_format3)
                    format_lengkap += 1
                    format_lengkap_array.append(str(i+1))
                if sumberReputasi(ref, 'conf'):
                    worksheet.write(row, 6, "Iya")
                    reputasi += 1
                    reputasi_array.append(str(i+1))
            elif 'Proc.' in listToString(ref['misc']):
                primer += 1
                primer_array.append(str(i+1))            
                worksheet.write(row, 3, "Iya")
                detail_format = cekConference(ref)
                if detail_format:
                    worksheet.write(row, 1, "Tidak Lengkap(" + listToString(detail_format) + ")", cell_format3)
                    format_lengkap += 1
                    format_lengkap_array.append(str(i+1))
                if sumberReputasi(ref, 'conf'):
                    worksheet.write(row, 6, "Iya")
                    reputasi += 1
                    reputasi_array.append(str(i+1))
            elif 'Procedia' in listToString(ref['misc']):
                primer += 1
                primer_array.append(str(i+1))            
                worksheet.write(row, 3, "Iya")
                detail_format = cekConference(ref)
                if detail_format:
                    worksheet.write(row, 1, "Tidak Lengkap(" + listToString(detail_format) + ")", cell_format3)
                    format_lengkap += 1
                    format_lengkap_array.append(str(i+1))
                if sumberReputasi(ref, 'conf'):
                    worksheet.write(row, 6, "Iya")
                    reputasi += 1
                    reputasi_array.append(str(i+1))
            elif 'Conf.' in listToString(ref['misc']):
                primer += 1
                primer_array.append(str(i+1))
                worksheet.write(row, 3, "Iya")
                detail_format = cekConference(ref)
                if detail_format:
                    worksheet.write(row, 1, "Tidak Lengkap(" + listToString(detail_format) + ")", cell_format3)
                    format_lengkap += 1
                    format_lengkap_array.append(str(i+1))
                if sumberReputasi(ref, 'conf'):
                    worksheet.write(row, 6, "Iya")
                    reputasi += 1
                    reputasi_array.append(str(i+1))
            elif 'Proceedings' in listToString(ref['misc']):
                primer += 1
                primer_array.append(str(i+1))
                worksheet.write(row, 3, "Iya")
                detail_format = cekConference(ref)
                if detail_format:
                    worksheet.write(row, 1, "Tidak Lengkap(" + listToString(detail_format) + ")", cell_format3)
                    format_lengkap += 1
                    format_lengkap_array.append(str(i+1))
                if sumberReputasi(ref, 'conf'):
                    worksheet.write(row, 6, "Iya")
                    reputasi += 1
                    reputasi_array.append(str(i+1))
            elif 'Konferensi' in listToString(ref['misc']):
                primer += 1
                primer_array.append(str(i+1))
                worksheet.write(row, 3, "Iya")
                detail_format = cekConference(ref)
                if detail_format:
                    worksheet.write(row, 1, "Tidak Lengkap(" + listToString(detail_format) + ")", cell_format3)
                    format_lengkap += 1
                    format_lengkap_array.append(str(i+1))
                if sumberReputasi(ref, 'conf'):
                    worksheet.write(row, 6, "Iya")
                    reputasi += 1
                    reputasi_array.append(str(i+1))
            elif 'Seminar' in listToString(ref['misc']):
                primer += 1
                primer_array.append(str(i+1))
                worksheet.write(row, 3, "Iya")
                detail_format = cekConference(ref)
                if detail_format:
                    worksheet.write(row, 1, "Tidak Lengkap(" + listToString(detail_format) + ")", cell_format3)
                    format_lengkap += 1
                    format_lengkap_array.append(str(i+1))
                if sumberReputasi(ref, 'conf'):
                    worksheet.write(row, 6, "Iya")
                    reputasi += 1
                    reputasi_array.append(str(i+1))
            elif 'Pros.' in listToString(ref['misc']):
                primer += 1
                primer_array.append(str(i+1))
                worksheet.write(row, 3, "Iya")
                detail_format = cekConference(ref)
                if detail_format:
                    worksheet.write(row, 1, "Tidak Lengkap(" + listToString(detail_format) + ")", cell_format3)
                    format_lengkap += 1
                    format_lengkap_array.append(str(i+1))
                if sumberReputasi(ref, 'conf'):
                    worksheet.write(row, 6, "Iya")
                    reputasi += 1
                    reputasi_array.append(str(i+1))

            #THESIS & DISSERTATION
            elif 'Thesis' in listToString(ref['misc']):
                primer += 1
                primer_array.append(str(i+1))
                worksheet.write(row, 3, "Iya")
                if not cekIsBuku(ref):
                    worksheet.write(row, 1, "Tidak Lengkap (*K)", cell_format3)
                    format_lengkap += 1
                    format_lengkap_array.append(str(i+1))
            elif 'thesis' in listToString(ref['misc']):
                primer += 1
                primer_array.append(str(i+1))
                worksheet.write(row, 3, "Iya")
                if not cekIsBuku(ref):
                    worksheet.write(row, 1, "Tidak Lengkap (*K)", cell_format3)
                    format_lengkap += 1
                    format_lengkap_array.append(str(i+1))
            elif 'Dissertation' in listToString(ref['misc']):
                primer += 1
                primer_array.append(str(i+1))
                worksheet.write(row, 3, "Iya")
                if not cekIsBuku(ref):
                    worksheet.write(row, 1, "Tidak Lengkap (*K)", cell_format3)
                    format_lengkap += 1
                    format_lengkap_array.append(str(i+1))
            elif 'dissertation' in listToString(ref['misc']):
                primer += 1
                primer_array.append(str(i+1))
                worksheet.write(row, 3, "Iya")
                if not cekIsBuku(ref):
                    worksheet.write(row, 1, "Tidak Lengkap (*K)", cell_format3)
                    format_lengkap += 1
                    format_lengkap_array.append(str(i+1))
                
            #LAIN LAIN
            elif 'vol. ' in listToString(ref['raw_ref']).lower():
                detail_format = cekJurnalIlmiah(ref)
                if detail_format:
                    worksheet.write(row, 1, "Tidak Lengkap(" + listToString(detail_format) + ")", cell_format3)
                    format_lengkap += 1
                    format_lengkap_array.append(str(i+1))
                primer += 1
                primer_array.append(str(i+1))
                worksheet.write(row, 3, "Iya")
                if sumberReputasi(ref, 'jurnal'):
                    worksheet.write(row, 6, "Iya")
                    reputasi += 1
                    reputasi_array.append(str(i+1))
                if sumberReputasi(ref, 'predatory'):
                    worksheet.write(row, 7, "Iya")
                    predatory += 1
                    predatory_array.append(str(i+1))
            #PP DI KOMEN SEMENTARA
            elif 'p. ' in listToString(ref['raw_ref']).lower():
                detail_format = cekJurnalIlmiah(ref)
                if detail_format:
                    worksheet.write(row, 1, "Tidak Lengkap(" + listToString(detail_format) + ")", cell_format3)
                    format_lengkap += 1
                    format_lengkap_array.append(str(i+1))
                primer += 1
                primer_array.append(str(i+1))
                worksheet.write(row, 3, "Iya")
                if sumberReputasi(ref, 'jurnal'):
                    worksheet.write(row, 6, "Iya")
                    reputasi += 1
                    reputasi_array.append(str(i+1))
                if sumberReputasi(ref, 'predatory'):
                    worksheet.write(row, 7, "Iya")
                    predatory += 1
                    predatory_array.append(str(i+1))
            elif 'no. ' in listToString(ref['raw_ref']).lower():
                detail_format = cekJurnalIlmiah(ref)
                if detail_format:
                    worksheet.write(row, 1, "Tidak Lengkap(" + listToString(detail_format) + ")", cell_format3)
                    format_lengkap += 1
                    format_lengkap_array.append(str(i+1))
                worksheet.write(row, 3, "Iya")
                primer += 1
                primer_array.append(str(i+1))
                if sumberReputasi(ref, 'jurnal'):
                    worksheet.write(row, 6, "Iya")
                    reputasi += 1
                    reputasi_array.append(str(i+1))
                if sumberReputasi(ref, 'predatory'):
                    worksheet.write(row, 7, "Iya")
                    predatory += 1
                    predatory_array.append(str(i+1))
            #BUKU
            elif cekIsBuku(ref):
                primer += 1
                primer_array.append(str(i+1))
                worksheet.write(row, 3, "Iya")
            else:
                cell_format2.set_bg_color('red')
                worksheet.write(row, 1, "Tidak Lengkap(*TT)", cell_format2)
                format_lengkap += 1
                format_lengkap_array.append(str(i+1))
        else:
            cell_format2.set_bg_color('red')
            worksheet.write(row, 1, "Tidak Lengkap(*TT)", cell_format2)
            format_lengkap += 1
            format_lengkap_array.append(str(i+1))
        row += 1


    #WRITE ACUAN PRIMER IN OUTPUT NARASI
    primer_array2 = []
    for i in range(primer_total):
        primer_array2.append(str(i+1))

    primer_array3 = []

    for primer_item in primer_array2:
        if primer_item not in primer_array:
            primer_array3.append(primer_item)

    primer = primer_total - primer
    primer_percentage = primer * 100 / primer_total
    primer_percentage_str = "%.2f" % primer_percentage
    primer_percentage_str2 = 100 - primer_percentage
    primer_percentage_str2 = "%.2f" % primer_percentage_str2

    cell_format_primer = workbook.add_format()
    cell_format_primer.set_border()
    cell_format_primer.set_text_wrap()

    if primer_percentage > 20:
        cell_format_primer.set_bg_color('red')
        narasi_primer = 'Referensi yang digunakan pada paper ini hanya menggunakan acuan primer sebesar ' + str(primer_percentage_str2) + '%. Referensi dianggap beracuan primer ketika referensi tersebut berasal dari Jurnal, Conference, Buku, Thesis, dan Disertasi. \n \nReferensi dengan Nomor ' + listToStringComma(primer_array3) +' tidak beracuan primer.'
        worksheet_narasi.write('A6', 'Acuan Primer', cell_format_primer)
        worksheet_narasi.write('B6', narasi_primer, cell_format_primer2)
    else:
        cell_format_primer.set_bg_color('#00ac56')
        if (len(primer_array3) > 0):
            narasi_primer = 'Referensi yang digunakan pada paper ini menggunakan acuan primer sebesar ' + str(primer_percentage_str2) +'%. Referensi dianggap beracuan primer ketika referensi tersebut berasal dari Jurnal, Conference, Buku, Thesis, dan Disertasi.  \n \nReferensi dengan Nomor ' + listToStringComma(primer_array3) +' tidak beracuan primer.'
        else:
            narasi_primer = 'Referensi yang digunakan pada paper ini menggunakan acuan primer sebesar ' + str(primer_percentage_str2) +'%. Referensi dianggap beracuan primer ketika referensi tersebut berasal dari Jurnal, Conference, Buku, Thesis, dan Disertasi.'
        worksheet_narasi.write('A6', 'Acuan Primer', cell_format_primer)
        worksheet_narasi.write('B6', narasi_primer, cell_format_primer2)

    #WRITE FORMAT LENGKAP IN OUTPUT NARASI

    cell_format_format = workbook.add_format()
    cell_format_format.set_border()
    cell_format_format.set_text_wrap()
    if format_lengkap > 0:
        cell_format_format.set_bg_color('red')
        narasi_format = 'Terdapat '+ str(format_lengkap) +' referensi yang memiliki format tidak lengkap. Referensi dianggap lengkap jika sudah mengikuti aturan referensi IEEE. Selain itu, referensi diluar Jurnal, Conference, Buku, Thesis, dan Disertasi juga akan dianggap tidak lengkap. \n \nReferensi dengan Nomor ' + listToStringComma(format_lengkap_array) + ' tidak berformat lengkap.'
        worksheet_narasi.write('A7', 'Format Lengkap', cell_format_format)
        worksheet_narasi.write('B7', narasi_format, cell_format_primer2)
    else:
        cell_format_format.set_bg_color('#00ac56')
        narasi_format = 'Format yang digunakan pada referensi ini sudah lengkap. Referensi dianggap lengkap jika sudah mengikuti aturan referensi IEEE. Selain itu, referensi diluar Jurnal, Conference, Buku, Thesis, dan Disertasi juga akan dianggap tidak lengkap'
        worksheet_narasi.write('A7', 'Format Lengkap', cell_format_format)
        worksheet_narasi.write('B7', narasi_format, cell_format_primer2)

    #WRITE TERSITASI IN OUTPUT NARASI
    cell_format_tersitasi = workbook.add_format()
    cell_format_tersitasi.set_border()
    cell_format_tersitasi.set_text_wrap()
    len_not_citated = len(not_citated)
    if len_not_citated > 0:
        cell_format_tersitasi.set_bg_color('red')
        narasi_format = 'Terdapat ' + str(len_not_citated) + ' referensi yang tidak tersitasi pada naskah. Referensi dianggap tidak tersitasi pada naskah ketika referensi pada artikel ilmiah sama sekali tidak disitasi pada keseluruhan artikel. \n \nReferensi dengan nomor ' + listToStringComma(not_citated) + ' tidak disitasi pada naskah'
        worksheet_narasi.write('A8', 'Tersitasi Pada Naskah', cell_format_tersitasi)
        worksheet_narasi.write('B8', narasi_format, cell_format_primer2)
    else:
        cell_format_tersitasi.set_bg_color('#00ac56')
        narasi_format = 'Seluruh referensi telah disitasi pada naskah minimal 1 kali. Referensi dianggap tidak tersitasi pada naskah ketika referensi pada artikel ilmiah sama sekali tidak disitasi pada keseluruhan artikel.'
        worksheet_narasi.write('A8', 'Tersitasi Pada Naskah', cell_format_tersitasi)
        worksheet_narasi.write('B8', narasi_format, cell_format_primer2)

    #WRITE SUMBER REPUTASI

    reputasi_array2 = []
    for i in range(reputasi_total):
        reputasi_array2.append(str(i+1))

    reputasi_array3 = []

    for reputasi_item in reputasi_array2:
        if reputasi_item not in reputasi_array:
            reputasi_array3.append(reputasi_item)

    cell_format_reputasi = workbook.add_format()
    cell_format_reputasi.set_border()
    cell_format_reputasi.set_text_wrap()

    reputasi = reputasi_total - reputasi
    reputasi_percentage = reputasi * 100 / reputasi_total
    reputasi_percentage_str = "%.2f" % reputasi_percentage
    reputasi_percentage_str2 = 100 - reputasi_percentage
    reputasi_percentage_str2 = "%.2f" % reputasi_percentage_str2

    if reputasi_percentage > 20:
        narasi_reputasi = 'Referensi yang digunakan pada pada paper ini masuk kedalam white list scopus sebesar ' + str(reputasi_percentage_str2) + '%.\n \nRefernsi dengan Nomor '+ listToStringComma(reputasi_array3) +' terindeks dalam white list scopus'
        worksheet_narasi.write('A9', 'White List', cell_format_reputasi)
        worksheet_narasi.write('B9', narasi_reputasi, cell_format_reputasi)
    else:
        if (len(reputasi_array3) > 0):
            narasi_reputasi = 'Referensi yang digunakan pada pada paper ini masuk kedalam white list scopus sebesar ' + str(reputasi_percentage_str2) +'%. \n \nRefernsi dengan Nomor '+ listToStringComma(reputasi_array3) +' terindeks dalam white list scopus'
        else:
            narasi_reputasi = 'Referensi yang digunakan pada pada paper ini masuk kedalam white list scopus sebesar ' + str(reputasi_percentage_str2) +'%.'
        worksheet_narasi.write('A9', 'White List', cell_format_reputasi)
        worksheet_narasi.write('B9', narasi_reputasi, cell_format_reputasi)
    


    #WRITE PREDATORY
    
    predatory_array2 = []
    for i in range(predatory_total):
        predatory_array2.append(str(i+1))

    predatory_array3 = []

    for predatory_item in predatory_array2:
        if predatory_item not in predatory_array:
            predatory_array3.append(predatory_item)

    cell_format_predatory = workbook.add_format()
    cell_format_predatory.set_border()
    cell_format_predatory.set_text_wrap()

    predatory = predatory_total - predatory
    predatory_percentage = predatory * 100 / predatory_total
    predatory_percentage_str = "%.2f" % predatory_percentage
    predatory_percentage_str2 = 100 - predatory_percentage
    predatory_percentage_str2 = "%.2f" % predatory_percentage_str2

    if predatory_percentage > 20:
        narasi_predatory = 'Referensi yang digunakan pada pada paper ini masuk kedalam black list sebesar ' + str(predatory_percentage_str2) + '%. Data black list yang digunakan berasal dari bealls list. \n \nReferensi dengan Nomor'+ listToStringComma(predatory_array3) +' terindeks dalam black list.' 
        worksheet_narasi.write('A10', 'Black List', cell_format_predatory)
        worksheet_narasi.write('B10', narasi_predatory, cell_format_predatory)
    else:
        if (len(predatory_array3) > 0):
            narasi_predatory = 'Referensi yang digunakan pada pada paper ini masuk kedalam black list sebesar ' + str(predatory_percentage_str2) +'%. Data black list yang digunakan berasal dari bealls list. \n \nReferensi dengan Nomor'+ listToStringComma(predatory_array3) +' terindeks dalam black list.'
        else:
            narasi_predatory = 'Referensi yang digunakan pada pada paper ini masuk kedalam black list sebesar ' + str(predatory_percentage_str2) +'%. Data black list yang digunakan berasal dari bealls list'
        worksheet_narasi.write('A10', 'Black List', cell_format_predatory)
        worksheet_narasi.write('B10', narasi_predatory, cell_format_predatory)

    workbook.close()

    if file_name != "":
        os.remove(os.path.join(settings.MEDIA_ROOT, file_name))

    if url_pdf != '':
        file_name_analisis = url_pdf + '_analisis.xlsx'
    else:
        file_name_analisis = file_pdf.name[:-4] + '_analisis.xlsx'

    BASE_DIR = os.path.dirname(os.path.dirname(os.path.abspath(__file__)))
    filepath = BASE_DIR + '/analisis.xlsx' 
    file_path = os.path.join(settings.MEDIA_ROOT, filepath)
    if os.path.exists(file_path):
        with open(file_path, 'rb') as fh:
            response = HttpResponse(fh.read(), content_type="application/vnd.ms-excel")
            response['Content-Disposition'] = 'inline; filename=' + os.path.basename(file_name_analisis)
            return response
    raise Http404
   